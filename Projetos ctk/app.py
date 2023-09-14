import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook


#aparencia do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()


    def layout_config(self):
        self.title("Sistema de Cadastro de Animais da Jotinha")
        self.geometry("700x600")


    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff', ]).place(x=50, y=530) 
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=560)
        


    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=800, height=50, corner_radius=0, bg_color="black", fg_color="black")
        frame.place(x=0, y=10)

        title = ctk.CTkLabel(frame, text="Sistema de Cadastro de Animais", font=("Century Gothic", 24), text_color="#fff").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do fomulário!", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path("Clientes.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            folha = ficheiro.active
            folha['A1']="Nome"
            folha['B1']="Contato"
            folha['C1']="Endereço"
            folha['D1']="Idade"
            folha['E1']="Sexo"
            folha['F1']="Castrado"
            folha['G1']="Email"
            folha['H1']="Quadro Clínico "

            ficheiro.save("Clientes.xlsx")

        def submit():
            

            #pegando os dados
            name = name_value.get()
            contato = contact_value.get()
            endereço = address_value.get()
            idade = age_value.get()
            sexo = sexo_combobox.get()
            esteril = esteril_combobox.get()
            email = email_value.get()
            obs  = obs_entry.get(0.0, END)

            if(name =="" or contato=="" or endereço == "" or idade =="" or obs==""):
                messagebox.showerror("Sistema", "ERRO! \n Digite todos os campos!")
            else:
                    
                ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                folha = ficheiro.active
                folha.cell(column = 1, row = folha.max_row+1, value = name )
                folha.cell(column = 2, row = folha.max_row, value = contato )
                folha.cell(column = 3, row = folha.max_row, value = endereço )
                folha.cell(column = 4, row = folha.max_row, value = idade )
                folha.cell(column = 5, row = folha.max_row, value = sexo )
                folha.cell(column = 6, row = folha.max_row, value = esteril )
                folha.cell(column = 7, row = folha.max_row, value = email )
                folha.cell(column = 8, row = folha.max_row, value = obs )

                ficheiro.save(r"Clientes.xlsx")
                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")


        def clear():
            name_value.set("")
            contact_value.set("")
            address_value.set("")
            age_value.set("")
            email_value.set("")
            obs_entry.delete(0.0,END)
            
            

        #texts variaveis
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        email_value = StringVar()
        

        #ENTRYS

        nome_entrada = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic", 16), fg_color="transparent").place(x=50, y=150)
        contato_entrada = ctk.CTkEntry(self, width=200, textvariable=contact_value,  font=("Century Gothic", 16), fg_color="transparent").place(x=450, y=150)
        endereço_entrada = ctk.CTkEntry(self, width=350, textvariable=address_value,  font=("Century Gothic", 16), fg_color="transparent").place(x=50, y=210)
        idade_entrada = ctk.CTkEntry(self, width=80, textvariable=age_value,  font=("Century Gothic", 16), fg_color="transparent").place(x=450, y=270)
        email_entrada = ctk.CTkEntry(self, width=350, textvariable=email_value,  font=("Century Gothic", 16), fg_color="transparent").place(x=50, y=270)

        #Combobox

        sexo_combobox = ctk.CTkComboBox(self, width= 90, height= 30, values=["Macho", "Fêmea"], font=("Century Gothic", 14))
        sexo_combobox.place(x=450, y=210)
        sexo_combobox.set("Macho")

        esteril_combobox = ctk.CTkComboBox(self, width= 90, height= 30, values=["Sim","Não"], font=("Century Gothic", 14))
        esteril_combobox.place(x=550, y=210)
       

        #TEXTBOX

        obs_entry = ctk.CTkTextbox(self, width= 500, height= 150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
        obs_entry.place(x=50, y=340)
        
        #LABELS

        nome = ctk.CTkLabel(self, text="Nome: ", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=50, y=120)

        contato = ctk.CTkLabel(self, text="Contato: ", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=450, y=120)

        endereço = ctk.CTkLabel(self, text="Endereço: ", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=50, y=180)

        idade = ctk.CTkLabel(self, text="Idade: ", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=450, y=240)

        email = ctk.CTkLabel(self, text="Email: ", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=50, y=240)

        sexo = ctk.CTkLabel(self, text="Sexo: ", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=450, y=180)

        esterilizada = ctk.CTkLabel(self, text="Castrado:", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=550, y=180)

        obs = ctk.CTkLabel(self, text="Quadro Clínico e Observações: ", font=("Century Gothic", 16), text_color=["#000", "#fff"]).place(x=50, y=310)

        #BOTOES

        btn_submeter = ctk.CTkButton(self, text="Salvar Dados".upper(), command=submit,  fg_color="#151", hover_color="#131").place(x=400, y=500)

        btn_submeter = ctk.CTkButton(self, text="Limpar Campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=210, y=500)

    
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)



if __name__ == "__main__":
    app = App()
    app.mainloop()